import pandas as pd
import numpy as np
from openpyxl import Workbook
from random import shuffle
import datetime

#oluşturulacak dosyanın açılması
wb = Workbook()
ws = wb.active

#dosyaları data frame e aktarma
df1 = pd.read_excel("list1.xlsx")
df2 = pd.read_excel("list2.xlsx")

#dosya düzenleme
df1["isim"] = df1["isim"].str.upper()
df2["isim"] = df2["isim"].str.upper()

#dataların aktarılacağı listeler
list1name = []
list2name = []



for i in range(len(df1["isim"])):
    #ilk liste için dataların satırlar olarak bir listeye alınıp listelerin oluşturulması
    row = df1.loc[i]
    list1name.append(row[1])


for i in range(len(df2["isim"])):
    #ikinci liste için data ayrıştırma
    row2 = df2.loc[i]
    list2name.append(row2[1])

#listelerin karıştırılması
shuffle(list1name)
shuffle(list2name)  
print(list1name)
print(list2name)

#liste uzunluklarının ayarlanması
length = len(list1name) + len(list2name)
print(length)


index1 = 0
index2 = 0

for i in range(length):
    if i%2 == 0:
        if index1 + 1 > len(list1name):
            #karıştırılan listenin girilmesi
            ws[f"B{i+1}"] = list2name[index2]
            #isme göre aratarak yeni bir dataframe oluşturma
            result2 =df2[df2["isim"].str.contains(f"{list2name[index2]}")][["koli"]]
            #yeni dataframein indexini alma
            list2 = result2.index.values.tolist()
            #indexle koli bilgisini alma
            koli2 = result2.loc[list2[0], "koli"]
            #veri girişi
            ws[f"C{i+1}"] = koli2
            ws[f"A{i+1}"] = i + 1
            #index artırma
            index2 = index2 + 1
        else:
            ws[f"B{i+1}"] = list1name[index1]
            result1 =df1[df1["isim"].str.contains(f"{list1name[index1]}")][["koli"]]
            list1 = result1.index.values.tolist()
            koli1 = result1.loc[list1[0], "koli"]
            ws[f"C{i+1}"] = koli1
            ws[f"A{i+1}"] = i + 1
            index1 = index1 + 1
    else:
        if index2 + 1 > len(list2name):
            ws[f"B{i+1}"] = list1name[index1]
            result1 =df1[df1["isim"].str.contains(f"{list1name[index1]}")][["koli"]]
            list1 = result1.index.values.tolist()
            koli1 = result1.loc[list1[0], "koli"]
            ws[f"C{i+1}"] = koli1
            ws[f"A{i+1}"] = i + 1
            index1 = index1 + 1
        else:  
            print(index2)     
            ws[f"B{i+1}"] = list2name[index2]
            result2 =df2[df2["isim"].str.contains(f"{list2name[index2]}")][["koli"]]
            list2 = result2.index.values.tolist()
            koli2 = result2.loc[list2[0], "koli"]
            ws[f"C{i+1}"] = koli2
            ws[f"A{i+1}"] = i + 1
            index2 = index2 + 1



today = datetime.datetime.today()

wb.save(f"{today.day}.{today.month}.{today.year} ürün teslim.xlsx")
